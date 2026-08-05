# -*- coding: utf-8 -*-
"""验收测试公共辅助模块：配置加载、登录、导入、导出、值规范化。"""
import json
import os
import requests

BASE = "http://localhost:8080"
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CFG_DIR = os.path.join(ROOT, "docs", "superpowers", "specs", "field-config")
ART = os.path.join(os.path.dirname(os.path.abspath(__file__)), "artifacts")
os.makedirs(ART, exist_ok=True)


def load_cfg(doc_type):
    p = os.path.join(CFG_DIR, doc_type + ".json")
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def login(username="admin", password="admin123"):
    r = requests.post(BASE + "/api/auth/login",
                      json={"username": username, "password": password}, timeout=10)
    assert r.status_code == 200, f"login http {r.status_code}: {r.text}"
    j = r.json()
    assert j.get("code") == 0, f"login err: {j}"
    return j["data"]["token"]


def hdr(token):
    return {"Authorization": "Bearer " + token, "Content-Type": "application/json"}


def api_list(token, doc_type, keyword=None, page=1, size=200):
    p = f"/api/doc/{doc_type}?page={page}&size={size}"
    if keyword:
        p += "&keyword=" + keyword
    r = requests.get(BASE + p, headers=hdr(token), timeout=15)
    j = r.json()
    assert j.get("code") == 0, f"list err: {j}"
    return j["data"]


def api_detail(token, doc_type, doc_id):
    r = requests.get(BASE + f"/api/doc/{doc_type}/{doc_id}", headers=hdr(token), timeout=10)
    j = r.json()
    assert j.get("code") == 0, f"detail err: {j}"
    return j["data"]


def api_import(token, doc_type, file_path):
    with open(file_path, "rb") as f:
        r = requests.post(BASE + f"/api/doc/{doc_type}/import",
                          files={"file": (os.path.basename(file_path), f,
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                          headers={"Authorization": "Bearer " + token}, timeout=120)
    j = r.json()
    return j.get("code"), j.get("message"), j.get("data")


def api_export(token, doc_type, keyword=None, out_path=None):
    p = f"/api/doc/{doc_type}/export"
    if keyword:
        p += "?keyword=" + keyword
    r = requests.get(BASE + p, headers={"Authorization": "Bearer " + token}, timeout=60)
    assert r.status_code == 200, f"export http {r.status_code}"
    if out_path is None:
        out_path = os.path.join(ART, f"{doc_type}_export.xlsx")
    with open(out_path, "wb") as f:
        f.write(r.content)
    cd = r.headers.get("Content-Disposition", "")
    return out_path, cd


def api_delete(token, doc_type, doc_id):
    r = requests.delete(BASE + f"/api/doc/{doc_type}/{doc_id}", headers=hdr(token), timeout=10)
    return r.json()


def norm(v):
    """将单元格值规范化为可比较的字符串（数字统一、去空白、None->''）。"""
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return repr(round(v, 10)).rstrip("0").rstrip(".")
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def read_sheet(path, sheet=0):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[sheet]
    return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]


def sheet_shape(path, sheet=0):
    rows = read_sheet(path, sheet)
    if not rows:
        return 0, 0
    maxc = max(len(r) for r in rows)
    return len(rows), maxc
